from flask import Flask,render_template,redirect,request,url_for,flash
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime
import flask_excel
from werkzeug.utils import secure_filename
from sqlalchemy import or_
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import session
from sqlalchemy import func
from forms import RegisterationForm,LoginForm,updateprofileForm,uploadfileForm,StudentaddForm,UpdateStudentForm,UnitaddForm,UpdateUnitForm,AbsenceForm,AbsenceFilterForm
from flask_bcrypt import Bcrypt
from flask_login import LoginManager,UserMixin,login_user,current_user,logout_user,login_required
from flask_wtf.file import FileField, FileRequired
import re
from wtforms_sqlalchemy.orm import model_form
from sqlalchemy import exc

app = Flask(__name__)
app.config['SECRET_KEY'] = 'zfgb6554vddfgsrg54351f' #for csrf token
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///Abschk2.db'  #name of db
db = SQLAlchemy(app)
bcrypt = Bcrypt(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'
login_manager.login_message = 'first login'

# for login and user mangement
class User(db.Model,UserMixin):
	id = db.Column(db.Integer, primary_key=True)
	username = db.Column(db.String(80), unique=True, nullable=False)
	email = db.Column(db.String(120), unique=True, nullable=False)
	password = db.Column(db.String(120), nullable=False)
	date = db.Column(db.DateTime, default=datetime.now)
	def __repr__(self):
		return f'User({self.id} - {self.username} - {self.email} - {self.password} - {self.date})'

# for absent record storage
class nopresense(db.Model):
	id = db.Column(db.Integer, primary_key=True);
	studentid = db.Column(db.Integer,unique=False, nullable=False);
	gname = db.Column(db.String(120), unique=False, nullable=False)
	surname = db.Column(db.String(120), unique=False, nullable=False)
	date = db.Column(db.Date, default=datetime.now)
	unitcode = db.Column(db.String(120), unique=False, nullable=False)
	classtype = db.Column(db.String(120), unique=False, nullable=False)
	reason = db.Column(db.String(180), unique=False, nullable=False)
	note = db.Column(db.String(280), unique=False, nullable=False)

# for unit data storage
class unit(db.Model):
	unitcode = db.Column(db.String(120), unique=True, nullable=False,primary_key=True)
	instructorname  = db.Column(db.String(120), unique=False, nullable=False)
	instructoremail = db.Column(db.String(120), unique=False, nullable=False)

# for student information storage
class Student(db.Model):
	id = db.Column(db.Integer,unique=True, nullable=False,primary_key=True);
	surname = db.Column(db.String(120), unique=False, nullable=False)
	title = db.Column(db.String(120), unique=False, nullable=False)
	gname = db.Column(db.String(120), unique=False, nullable=False)
	teachperiod = db.Column(db.String(120), unique=False, nullable=False)
	unitcode = db.Column(db.String(120), unique=False, nullable=False)
	unitmode = db.Column(db.String(120), unique=False, nullable=False)
	unitstatus = db.Column(db.String(120), unique=False, nullable=False)
	crsstatus = db.Column(db.String(120), unique=False, nullable=False)
	crsscode = db.Column(db.String(120), unique=False, nullable=False)#
	email = db.Column(db.String(120), unique=False, nullable=False)
	unittitle = db.Column(db.String(120), unique=False, nullable=False)
	pgname = db.Column(db.String(120), unique=False, nullable=True)
	# unitstatus = db.Column(db.String(120), unique=True, nullable=False) course code


regex = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'

def check(email):
    if(re.fullmatch(regex, email)):
        return True
    else:
    	return False








# safhe user 
@login_manager.user_loader
def load_user(user_id):
	return User.query.get(int(user_id))



# upload file student data
@app.route('/manulaletter', methods=['GET','POST'])
@login_required
def manulaletter():
	if current_user.username!='admin':
		return render_template('error.html')
	table = " [.....] "
	if request.method=='GET':
		return render_template("manulaletter.html" ,table=table )
	if request.method=="POST":
		return render_template("manulaletter.html" ,table=table )
		f = open("letter.txt", "a")
		f.write("Dear[Nominated course coordinator],")
		f.write("\n")
		f.write("The followings are the student(s) absence details:")
		f.write(table+"Please review their absences on the following report before taking appropriate action.")
		f.write("For assistance, please contact the PMSD ASO.\n\n")
		f.close()




# upload file student data
@app.route('/upload', methods=['GET','POST'])
@login_required
def upload_file():
	if current_user.username!='admin':
		return render_template('error.html')
	form = uploadfileForm()
	if request.method=='GET':
		return render_template("upload.html" , form=form)
	if request.method=="POST":
		if not form.validate_on_submit():
			flash(form.errors)
			return render_template('upload.html',form=form)
		filename = secure_filename("xlsxfileuploaded.xlsx")
		# filename = secure_filename(form.xlsxfile.data.filename)
		form.xlsxfile.data.save('uploads/' + filename)
		return redirect(url_for('parsexlsxfile'))

# student detail store
@app.route('/parsexlsxfile', methods=['GET','POST'])
@login_required
def parsexlsxfile():
	loc = ("uploads/xlsxfileuploaded.xlsx") 
	wb = load_workbook(loc) 
	sheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])
	i = 1;
	while (str(sheet['A'+str(i)].value)!="None"):
		i = i+1
		# addtosbxlsx(i)
		try:
			q = unit.query.filter(unit.unitcode == str(sheet['H'+str(i)].value)).count()
			if q:
				st = Student(id = str(sheet['A'+str(i)].value),surname = str(sheet['B'+str(i)].value),title = str(sheet['C'+str(i)].value),gname = str(sheet['D'+str(i)].value),teachperiod = str(sheet['G'+str(i)].value),unitcode = str(sheet['H'+str(i)].value),unitmode = str(sheet['J'+str(i)].value),unitstatus = str(sheet['L'+str(i)].value),crsstatus = str(sheet['S'+str(i)].value),email = str(sheet['AF'+str(i)].value),unittitle = str(sheet['AG'+str(i)].value),pgname = str(sheet['AH'+str(i)].value),crsscode=str(sheet['O'+str(i)].value))
				db.session.add(st)
				db.session.commit()
		except exc.SQLAlchemyError as e:
			# if i > 5:
			# 	return str(e)
			db.session.rollback()
			pass
	# return str(i);	
	return redirect(url_for('student'))

def addtosbxlsx(i):
	loc = ("uploads/xlsxfileuploaded.xlsx") 
	wb = load_workbook(loc) 
	sheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])	
	try:
		st = Student(id = str(sheet['A'+str(i)].value),surname = str(sheet['B'+str(i)].value),title = str(sheet['C'+str(i)].value),gname = str(sheet['D'+str(i)].value),teachperiod = str(sheet['G'+str(i)].value),unitcode = str(sheet['H'+str(i)].value),unitmode = str(sheet['J'+str(i)].value),unitstatus = str(sheet['L'+str(i)].value),crsstatus = str(sheet['S'+str(i)].value),email = str(sheet['AF'+str(i)].value),unittitle = str(sheet['AG'+str(i)].value),pgname = str(sheet['AH'+str(i)].value),crsscode=str(sheet['O'+str(i)].value))
		db.session.add(st)
		db.session.commit()
	except:
		a = i


# upload file abscent 
@app.route('/upload2', methods=['GET','POST'])
@login_required
def upload2_file():
	if current_user.username!='admin':
		return render_template('error.html')
	form = uploadfileForm()
	if request.method=='GET':
		return render_template("upload2.html" , form=form)
	if request.method=="POST":
		if not form.validate_on_submit():
			flash(form.errors)
			return render_template('upload2.html',form=form)		
		filename = secure_filename("xlsxfileuploaded2.xlsx")
		# filename = secure_filename(form.xlsxfile.data.filename)
		form.xlsxfile.data.save('uploads/' + filename)
		return redirect(url_for('parsexlsxfile2'))

# absecnt data storage
@app.route('/parsexlsxfile2', methods=['GET','POST'])
@login_required
def parsexlsxfile2():
	loc = ("uploads/xlsxfileuploaded2.xlsx") 
	wb = load_workbook(loc) 
	sheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])
	# sheet = wb.sheet_by_index(0)
	i = 2;
	# return "--"+str(sheet['A25'].value)+"--"
	while (str(sheet['A'+str(i)].value)!="None"):
		# st = nopresense(id = str(sheet['A'+str(i)].value),surname = str(sheet['B'+str(i)].value),title = str(sheet['C'+str(i)].value),gname =datetime.datetime.strptime(str(sheet['D'+str(i)].value), '%m %d %Y').date() str(sheet['D'+str(i)].value),teachperiod = str(sheet['G'+str(i)].value),unitcode = str(sheet['H'+str(i)].value),unitmode = str(sheet['J'+str(i)].value),unitstatus = str(sheet['L'+str(i)].value),crsstatus = str(sheet['S'+str(i)].value),email = str(sheet['A'+str(i)].value),unittitle = str(sheet['A'+str(i)].value),pgname = str(sheet['D'+str(i)].value))
		try:
			q = unit.query.filter(unit.unitcode == str(sheet['E'+str(i)].value)).count()
			q1 = Student.query.filter(Student.id == str(sheet['A'+str(i)].value)).count()
			q2 = Student.query.filter(Student.gname == str(sheet['B'+str(i)].value)).count()
			q3 = Student.query.filter(Student.surname == str(sheet['C'+str(i)].value)).count()
			if q and q1>0 and q2>0 and q3>0:
				st = nopresense(studentid = str(sheet['A'+str(i)].value),gname = str(sheet['B'+str(i)].value),surname = str(sheet['C'+str(i)].value),date =datetime.strptime(str(sheet['D'+str(i)].value), '%Y-%m-%d  %H:%M:%S').date(),unitcode = str(sheet['E'+str(i)].value),classtype = str(sheet['F'+str(i)].value),reason = str(sheet['G'+str(i)].value),note =str(sheet['H'+str(i)].value))
				db.session.add(st)
				db.session.commit()
		except Exception as e:
			db.session.rollback()
			pass	
		i = i+1
	return redirect(url_for('absencemng'))

# student page
@app.route('/student', methods=['GET','POST'])
@login_required
def student():
	if current_user.username!='admin':
		return render_template('error.html')
	form = StudentaddForm()
	if request.method=="POST":
		# if not form.validate_on_submit():
		# 	flash(form.errors)
		# 	return render_template('student.html',form=form)
		if not check(form.email.data):
			flash("wrong email")
			return render_template('student.html',form=form)
		q = unit.query.filter(unit.unitcode == form.unitcode.data).count()
		# q = unit.query(unit.unitcode).filter(unit.unitcode==form.unitcode.data)
		if not (q>0):
			flash("unit code not found ")
			return render_template('student.html',form=form)
		try:
			st = Student(id = form.id.data,surname = form.surname.data,title = form.title.data,gname = form.gname.data,teachperiod = form.teachperiod.data,unitcode = form.unitcode.data,unitmode = form.unitmode.data,unitstatus = form.unitstatus.data,crsstatus = form.crsstatus.data,email = form.email.data,unittitle = form.unittitle.data,pgname = form.pgname.data,crsscode=form.crsscode.data)
			db.session.add(st)
			db.session.commit()
			flash("student added successfully")
		except Exception as e:
			flash("something went wring(check duplicate id)")
		return redirect(url_for('student'))
	# return redirect('/success')
	if request.method=='GET':
		groups_list= [g.unitcode for g in unit.query.all()]
		form.unitcode.choices = groups_list
		return render_template('student.html' , form=form)  

# show list of students
@app.route('/studentshow', methods=['GET','POST'])
@login_required
def studentshow():
	us = Student.query.all();
	return render_template('student_show.html',users=us)
	  
# delete student	  # 
@app.route('/deletestudent/<stud_id>')
@login_required
def deletestudent(stud_id):
	if current_user.username!='admin':
		return render_template('error.html')
	Student.query.filter(Student.id == stud_id).delete()
	db.session.commit()	
	return redirect(url_for('studentshow'))	  

# edit student data
@app.route('/student/<stud_id>', methods=('GET', 'POST'))
@login_required
def detailstudent(stud_id):
	# return "dfgf"
	stndt = Student.query.get(stud_id)
	# return str(stndt.id)
	form = UpdateStudentForm()	
	if request.method=="POST":		
		if current_user.username!='admin':
			return render_template('error.html')
		if not form.validate_on_submit():
			# return str(form.errors)
			if "{'unitcode': ['Not a valid choice']}" != str(form.errors):
				flash(form.errors)
				return render_template('one_student.html',form=form)		
		if not check(form.email.data):
			flash("wrong email")
			return render_template('one_student.html',form=form)			
		stndt.id = form.id.data 
		stndt.surname = form.surname.data 
		stndt.title = form.title.data 
		stndt.gname = form.gname.data 
		stndt.teachperiod = form.teachperiod.data 
		stndt.unitcode = form.unitcode.data 
		stndt.unitmode = form.unitmode.data 
		stndt.unitstatus = form.unitstatus.data 
		stndt.crsstatus = form.crsstatus.data 
		stndt.email = form.email.data 
		stndt.unittitle = form.unittitle.data 
		stndt.pgname = form.pgname.data 
		stndt.crsscode = form.crsscode.data
		db.session.commit()
		flash("update successfully")
		return redirect(url_for('student'))	
	if request.method=="GET":	
		groups_list= [g.unitcode for g in unit.query.all()]
		form.unitcode.choices = groups_list
		form.id.data = stndt.id 
		form.surname.data = stndt.surname 
		form.title.data = stndt.title 
		form.gname.data = stndt.gname 
		form.teachperiod.data = stndt.teachperiod 
		form.unitcode.data = stndt.unitcode 
		form.unitmode.data = stndt.unitmode 
		form.unitstatus.data = stndt.unitstatus 
		form.crsstatus.data = stndt.crsstatus 
		form.email.data = stndt.email 
		form.unittitle.data = stndt.unittitle 
		form.pgname.data = stndt.pgname 
		form.crsscode.data = stndt.crsscode 
		return render_template('one_student.html' , user = stndt , form=form)

@app.route('/unitt', methods=['GET','POST'])
@login_required
def unitt():
	if current_user.username!='admin':
		return render_template('error.html')
	form = UnitaddForm()
	if request.method=="POST":
		if not form.validate_on_submit():
			flash(form.errors)
			return render_template('unit.html',form=form)	
		if not check(form.instructoremail.data):
			flash("wrong email")
			return render_template('unit.html',form=form)				
		# return redirect('/success')
		try:
			unt = unit(unitcode = form.unitcode.data,instructorname = form.instructorname.data,instructoremail = form.instructoremail.data)
			db.session.add(unt)
			db.session.commit()
			flash("UNIT added successfully")
		except Exception as e:
			flash("something went wring(check duplicate unitcode)")
		return redirect(url_for('unitt'))
	# return redirect('/success')
	if request.method=='GET':
		return render_template('unit.html' , form=form)  

@app.route('/unitsshow', methods=['GET','POST'])
@login_required
def unitsshow():
	us = unit.query.all();
	return render_template('units_show.html',users=us)

@app.route('/deleteunit/<unit_id>')
@login_required
def deleteunit(unit_id):
	if current_user.username!='admin':
		return render_template('error.html')
	unit.query.filter(unit.unitcode == unit_id).delete()
	db.session.commit()	
	return redirect(url_for('unitsshow'))	

@app.route('/unitt/<unit_id>', methods=('GET', 'POST'))
@login_required
def detailunit(unit_id):
	# return "dfgf"
	unt = unit.query.get(unit_id)
	# return stndt.id.data
	form = UpdateUnitForm()	
	if request.method=="POST":
		if current_user.username!='admin':
			return render_template('error.html')
		if not form.validate_on_submit():
			flash(form.errors)
			return render_template('one_unit.html',form=form)			
		# return user
		unt.unitcode = form.unitcode.data 
		unt.instructorname = form.instructorname.data 
		unt.instructoremail = form.instructoremail.data 
		db.session.commit()
		flash("update successfully")
		return redirect(url_for('unitt'))	
	if request.method=="GET":	
		form.unitcode.data = unt.unitcode 
		form.instructorname.data = unt.instructorname 
		form.instructoremail.data = unt.instructoremail 
		return render_template('one_unit.html' , user = unt , form=form)

@app.route('/absencemng', methods=['GET','POST'])
@login_required
def absencemng():
	if current_user.username!='admin':
		return redirect(url_for('checkabsence'))
	form = AbsenceForm()
	if request.method=="POST":
		# if not form.validate_on_submit():
		# 	flash(form.errors)
		# 	return render_template('absence.html',form=form)		
		# return redirect('/success')
		q1 = Student.query.filter(Student.id ==  form.studentid.data).count()
		if q1<1:
			flash("Student Id not found")
			return render_template('absence.html',form=form)
		abs = nopresense(studentid = form.studentid.data,gname = form.gname.data,surname = form.surname.data,date = form.date.data,unitcode = form.unitcode.data,classtype = form.classtype.data,reason = form.reason.data,note = form.note.data)
		db.session.add(abs)
		db.session.commit()
		flash("record added successfully")
		return redirect(url_for('absencemng'))
	# return redirect('/success')
	if request.method=='GET':
		groups_list= [g.unitcode for g in unit.query.all()]
		form.unitcode.choices = groups_list
		return render_template('absence.html' , form=form)  

@app.route('/absenceshow', methods=['GET','POST'])
@login_required
def absenceshow():
	if current_user.username!='admin':
		return render_template('error.html')	
	us = nopresense.query.all();
	return render_template('absence_show.html',users=us)

@app.route('/deleterecord/<abscid>')
@login_required
def deleterecord(abscid):
	if current_user.username!='admin':
		return render_template('error.html')
	nopresense.query.filter(nopresense.id == abscid).delete()
	db.session.commit()	
	return redirect(url_for('absenceshow'))	

@app.route('/absencerecord/<abscid>', methods=('GET', 'POST'))
@login_required
def absencerecord(abscid):
	unt = nopresense.query.get(abscid)
	form = AbsenceForm()	
	if request.method=="POST":
		# return user
		# unt.id = form.id.data
		if current_user.username!='admin':
			return render_template('error.html')
		# if not form.validate_on_submit():
		# 	flash(form.errors)
		# 	return render_template('one_record.html',form=form)
		unt.studentid = form.studentid.data
		unt.gname = form.gname.data
		unt.surname = form.surname.data
		unt.date = form.date.data
		unt.unitcode = form.unitcode.data
		unt.classtype = form.classtype.data
		unt.reason = form.reason.data
		unt.note = form.note.data
		db.session.commit()
		flash("update successfully")
		return redirect(url_for('absenceshow'))	
	if request.method=="GET":	
		# form.id.data = unt.id
		form.studentid.data = unt.studentid
		form.gname.data = unt.gname
		form.surname.data = unt.surname
		# print (unt)
		# return unt.date.strftime("%Y-%m-%d %H:%M")

		groups_list= [g.unitcode for g in unit.query.all()]
		form.unitcode.choices = groups_list
		
		form.date.data = unt.date
		dd = "date: "+unt.date.strftime('%Y') + "/" + unt.date.strftime('%M') + "/" + unt.date.strftime('%d') + "(set new if you want)"
		form.unitcode.data = unt.unitcode
		form.classtype.data = unt.classtype
		form.reason.data = unt.reason
		form.note.data = unt.note 
		return render_template('one_record.html' , user = unt , dd=dd, form=form)



@app.route('/checkabsence', methods=['GET','POST'])
@login_required
def checkabsence():
	form = AbsenceFilterForm()
	# us = nopresense.query.all();
	# return render_template('checkabsence.html',users=us)
	if request.method=="POST":
		# if not form.validate_on_submit():
		# 	flash(form.errors)
		# 	return render_template('checkabsence.html',form=form)		
	# unt = nopresense.query.filter(or_(nopresense.studentid==form.studentid.data,nopresense.gname==form.gname.data,nopresense.surname==form.surname.data,nopresense.date==form.date.data,nopresense.unitcode==form.unitcode.data,nopresense.classtype==form.classtype))
	# unt = nopresense.query.filter(or_(nopresense.classtype==form.classtype))
	# unt = nopresense.query.filter(nopresense.studentid == form.studentid)
	# unt = nopresense.query.all()
		# unt = session.query(nopresense).all()
		unt = nopresense.query
		if form.classtype.data:
			unt = unt.filter(nopresense.classtype == form.classtype.data);
		if form.unitcode.data:
			unt = unt.filter(nopresense.unitcode == form.unitcode.data);
		if form.datefrom.data:
			unt = unt.filter(nopresense.date >=form.datefrom.data);
		if form.dateto.data:
			unt = unt.filter(nopresense.date <= form.dateto.data);
		if form.surname.data:
			unt = unt.filter(nopresense.surname == form.surname.data);
		if form.gname.data:
			unt = unt.filter(nopresense.gname == form.gname.data);
		if form.studentid.data:
			unt = unt.filter(nopresense.studentid == form.studentid.data);	
		# results = query.all()
		# return  str(form.treshhold.data)
		
		# unt.group_by(nopresense.studentid).having(func.count(nopresense.surname)>form.treshhold.data).all()
		unt.order_by(nopresense.date.desc())
		freq = {}
		for user in unt:
			if (user.studentid in freq):
				freq[user.studentid] += 1
			else:
				freq[user.studentid] = 1
		# return freq
		if form.publish.data:
			return render_template('manulaletter.html',fr =form.datefrom.data ,to =form.dateto.data   , users=unt , trsh =request.form.get('treshhold', type=int)  , freq=freq, form=form)
		return render_template('checkabsence.html' ,fr =form.datefrom.data ,to =form.dateto.data   , users=unt , trsh =request.form.get('treshhold', type=int)  , freq=freq, form=form)

	if request.method=="GET":	
		return render_template('checkabsence.html' , form=form)

@app.route('/sendmail/<idd>/<no>/<fr>/<en>')
@login_required
def sendmail(idd,no,fr,en):
	# if current_user.username!='admin':
	# 	return render_template('error.html')
	nprs = nopresense.query.get(idd)
	# return str(nprs.studentid)

	stdnt = Student.query.get(nprs.studentid)
	# return str(stdnt)
	unity = unit.query.get(nprs.unitcode)
	# instructoremail = unity.instructoremail
	# instructorename = unity.instructorname
	# studentid = stdnt.id
	# studentname = stdnt.gname
	number = no
	fr = fr  
	en = en
	f = open("mail.txt", "a")
	f.write(unity.instructoremail)
	f.write("\n")
	f.write("Dear "+unity.instructorname+"\n"+stdnt.gname+"["+str(stdnt.id)+"] has been reported as absent on ["+str(number)+"] different days between[ from: "+fr+ ", to:"+en+"]\n") 
	f.write("Please review their absences on [http://127.0.0.1:5000/checkabsence] before taking appropriate action.")
	f.write("For assistance, please contact the PMSD ASO.\n\n")
	f.close()
	return render_template('sendmail.html',number=number ,fr=fr , en=en , unity = unity , stdnt=stdnt)

# ////////////////////////////////////////////////////////
# ////////////////////////////////////////////////////////
# ////////////////////////////////////////////////////////
# ////////////////////////////////////////////////////////
# ////////////////////////////////////////////////////////
# ////////////////////////////////////////////////////////
# safhe asli
@app.route('/')
def intro():
    return render_template('intro.html')

@app.route('/login',methods=['GET','POST'])
def login():
	if current_user.is_authenticated:
		flash("you logged in before")
		return redirect(url_for('intro'))
	form = LoginForm()
	if request.method=='GET':
		return render_template('login.html' , form=form)
	if request.method=="POST":#check kardan method ersal etelaat
		userr = User.query.filter_by(email=form.email.data).first()
		if userr and bcrypt.check_password_hash( userr.password ,form.password.data):
			login_user(userr,remember=form.remember.data)
			flash("you logged in successfully")#namayesh payam be soorat flash 
			return redirect(url_for('intro')) 
		else:
			flash("email or password is incorrect")
			return redirect(url_for('intro'))

# register karadan karbar

@app.route('/register', methods=('GET', 'POST'))
@login_required
def register():
	form = RegisterationForm()
	if request.method=="POST":
		if not form.validate_on_submit():
			flash(form.errors)
			return render_template('register.html',form=form)	
		if not check(form.email.data):
			flash("wrong email")
			return render_template('register.html',form=form)
		# return redirect('/success')
		hashedpass = bcrypt.generate_password_hash(form.password.data).decode('utf-8')
		us = User(username=form.username.data,email=form.email.data,password=hashedpass)
		db.session.add(us)
		db.session.commit()
		flash("user registerd successfully")
		return redirect(url_for('intro'))
	# return redirect('/success')
	if request.method=='GET':
		return render_template('register.html' , form=form)

# ////////////////////////////////////////////////////////
# ////////////////////////////////////////////////////////
# ////////////////////////////////////////////////////////
# ////////////////////////////////////////////////////////
# ////////////////////////////////////////////////////////
# ////////////////////////////////////////////////////////

# ----------------------- routes



@app.route('/setting')
@login_required
def setting():
	if current_user.username!='admin':
		return render_template('error.html')
	us = User.query.all();

	return render_template('setting.html',users=us)    



@app.route('/deleteuser/<user_id>')
@login_required
def deleteuser(user_id):
	if current_user.username!='admin':
		return render_template('error.html')
	User.query.filter(User.id == user_id).delete()
	db.session.commit()	
	return redirect(url_for('setting'))

# ///////////////////////???
@app.route('/profile')
@login_required
def profile():
	# games = Game.query.filter_by(id=current_user.id)
	gameses = current_user.games
	quizess = current_user.queezes
	minn = 1000;
	maxx = 0;
	total = 0;
	for game in gameses:
		if(game.score<minn):
			minn = game.score
		if(game.score>maxx):
			maxx = game.score			
		total = total +1			
	return render_template('profile.html',gamess=gameses,quizes=quizess,maxx = maxx, minn = minn , total=total)    






@app.route('/user/<user_id>', methods=('GET', 'POST'))
def detail(user_id):
	# return "dfgf"
	user = User.query.get(user_id)
	form = updateprofileForm()	
	if request.method=="POST":
		# return user
		user.username = form.username.data
		user.email = form.email.data
		user.password =  bcrypt.generate_password_hash(form.password.data).decode('utf-8')
		db.session.commit()
		flash("update successfully")
		return redirect(url_for('setting'))
	if request.method=="GET":	
		form.username.data = user.username
		form.email.data = user.email
		form.password.data = user.password
		return render_template('user.html' , user = user , form=form)





@app.route('/logout')
def logout():
	if current_user.is_authenticated:
		logout_user()
		flash("Logout successful ")
	else:
		flash("you are not logged in")
	return redirect(url_for('intro'))






if __name__ == '__main__':
    app.run(debug=True)